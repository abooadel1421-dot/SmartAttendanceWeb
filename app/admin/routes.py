# app/admin/routes.py  
import time  
import threading  
import uuid # لاستخدامها في session_id  
# تأكد أن لديك datetime, timedelta, date مستوردة بالفعل  
from datetime import datetime, timedelta, date  
from flask import jsonify, request, session, flash, redirect, url_for, render_template  
from flask_login import login_required, current_user  
from .. import db  
from app.models.user import User, UserRole  
from app.models.student import Student  
from app.models.card import Card, CardStatus # تأكد من استيراد CardStatus  
from app.models.device import Device  
from app.models.attendance_log import AttendanceLog  
from app.admin.forms import UserForm, StudentForm, CardForm, DeviceForm  
from functools import wraps  
import pytz  
from sqlalchemy.orm import joinedload  
from . import admin_bp  


import csv
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from datetime import datetime as dt
import pytz
# ... (باقي الإستيرادات الموجودة لديك)
# app/admin/routes.py

import functools # 🟢 هذا مهم لوظيفة admin_required

from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, session, current_app
from flask_login import login_required, current_user
from sqlalchemy.orm import joinedload
from sqlalchemy import func
from app import db 
from app.models.user import User, UserRole 
from app.models.student import Student
from app.models.device import Device 
from app.models.card import Card 

# 🟢 تأكد من استيراد AttendanceLog و AttendanceStatus
from app.models.attendance_log import AttendanceLog, AttendanceStatus 

# 🟢 تأكد من استيراد AttendanceSummary
from app.models.attendance_summary import AttendanceSummary 

# 🟢 تأكد من استيراد datetime و date و time و timedelta
from datetime import datetime, date, time, timedelta 
import pytz
import io # 🟢 لاستخدام BytesIO في تصدير الـ PDF

# ... (بقية الكود) ...

latest_scanned_card = {  
    'uid': None,  
    'timestamp': None,  
    'session_id': None  # لربط القراءة بجلسة محددة  
}  

# قفل للحماية من التداخل  
scan_lock = threading.Lock()  


# Decorator to restrict access to admin users only  
def admin_required(f):  
    @wraps(f)  
    def decorated_function(*args, **kwargs):  
        print(f"\n--- DEBUG: Inside admin_required for function: {f.__name__} ---")  
        print(f"DEBUG: Is user authenticated? {current_user.is_authenticated}")  

        if current_user.is_authenticated:  
            print(f"DEBUG: Current user username: {current_user.username}")  
            print(f"DEBUG: Current user role: {current_user.role}")  
            print(f"DEBUG: Type of current_user.role: {type(current_user.role)}")  
            print(f"DEBUG: Comparison result (current_user.role == UserRole.ADMIN): {current_user.role == UserRole.ADMIN}")  
        else:  
            print("DEBUG: User is not authenticated. Redirecting to login.")  

        if not current_user.is_authenticated or current_user.role != UserRole.ADMIN:  
            flash('غير مصرح لك بالوصول إلى هذه الصفحة.', 'danger')  
            if current_user.is_authenticated and current_user.role == UserRole.TEACHER:  
                return redirect(url_for('teacher.dashboard')) # <--- تم التصحيح هنا: من teacher_bp إلى teacher  
            elif current_user.is_authenticated and current_user.role == UserRole.STUDENT:  
                return redirect(url_for('main.dashboard')) # هذا السطر صحيح بالفعل  
            else:  
                return redirect(url_for('main.index'))  

        print("DEBUG: User is authorized as admin. Proceeding to function.")  
        return f(*args, **kwargs)  
    return decorated_function  

@admin_bp.route('/')  
@login_required  
@admin_required  
def index():  
    total_users = User.query.count()  
    total_students = Student.query.count()  
    active_cards = Card.query.filter_by(status='ACTIVE').count()  
    total_devices = Device.query.count()  
    total_admins = User.query.filter_by(role=UserRole.ADMIN).count()  

    today_utc = datetime.now(pytz.utc).date()  
    attendance_today = AttendanceLog.query.filter(  
        db.func.date(AttendanceLog.timestamp) == today_utc  
    ).count()  

    seven_days_ago_utc = datetime.now(pytz.utc) - timedelta(days=7)  
    recent_students = Student.query.filter(Student.created_at >= seven_days_ago_utc).order_by(Student.created_at.desc()).limit(5).all()  

    latest_attendance_logs = AttendanceLog.query.order_by(AttendanceLog.timestamp.desc()).limit(5).all()  

    return render_template('admin/index.html',  
                           title='لوحة تحكم المسؤول',  
                           total_users=total_users,  
                           total_students=total_students,  
                           active_cards=active_cards,  
                           total_devices=total_devices,  
                           attendance_today=attendance_today,  
                           total_admins=total_admins,  
                           recent_students=recent_students,  
                           latest_attendance_logs=latest_attendance_logs,  
                           active_page='admin_dashboard',  
                           active_menu='dashboard')  

# --- إدارة المستخدمين (Users Management) ---  

@admin_bp.route('/users')  
@login_required  
@admin_required  
def manage_users():  
    users = User.query.all()  
    return render_template('admin/manage_users.html',  
                           title='إدارة المستخدمين',  
                           users=users,  
                           active_page='manage_users',  
                           active_menu='users')  

@admin_bp.route('/user/add', methods=['GET', 'POST'])  
@login_required  
@admin_required  
def add_user():  
    form = UserForm()  
    # لتعبئة قائمة الطلاب غير المرتبطين في النموذج (إذا كان موجودًا)  
    # هذا يفترض أن لديك حقل SelectField في UserForm يسمى 'student_to_link'  
    # ويجب أن يكون هذا الحقل متاحًا فقط إذا تم اختيار دور الطالب  
    # للتبسيط، سنقوم بتعبئته هنا، ولكن قد تحتاج إلى JavaScript في الواجهة الأمامية لإخفائه/إظهاره  
    unlinked_students = Student.query.filter(Student.user_account == None).order_by(Student.first_name).all()  
    form.student_to_link.choices = [(s.id, s.full_name) for s in unlinked_students]  
    form.student_to_link.choices.insert(0, ('', '--- اختر طالب للربط (اختياري) ---')) # قيمة فارغة لعدم الربط  

    if form.validate_on_submit():  
        user = User(username=form.username.data, email=form.email.data, role=form.role.data)  
        user.set_password(form.password.data)  

        # ****** الجزء الجديد للربط عند إضافة المستخدم ******  
        if user.role == UserRole.STUDENT:  
            # إذا اختار المسؤول طالبًا لربطه  
            if form.student_to_link.data:  
                student_id_to_link = int(form.student_to_link.data)  
                student_to_assign = Student.query.get(student_id_to_link)  

                if student_to_assign:  
                    # تحقق إذا كان الطالب مرتبطًا بالفعل بمستخدم آخر  
                    if student_to_assign.user_account:  
                        flash(f'الطالب {student_to_assign.full_name} مرتبط بالفعل بالمستخدم {student_to_assign.user_account.username}.', 'danger')  
                        return render_template('admin/user_form.html', title='إضافة مستخدم جديد', form=form, active_page='add_user', active_menu='users')  
                    user.student_profile = student_to_assign # ربط المستخدم بالطالب  
                else:  
                    flash('الطالب المحدد غير موجود.', 'danger')  
                    return render_template('admin/user_form.html', title='إضافة مستخدم جديد', form=form, active_page='add_user', active_menu='users')  
            # إذا لم يتم اختيار طالب، سيتم إنشاء المستخدم كطالب بدون ملف طالب مرتبط في البداية  
            # وهذا مقبول لأن student_id في User هو nullable=True  
        # **************************************************  

        db.session.add(user)  
        db.session.commit()  
        flash('تم إضافة المستخدم بنجاح!', 'success')  
        return redirect(url_for('admin.manage_users')) # تأكد من اسم الـ blueprint  
    return render_template('admin/user_form.html',  
                           title='إضافة مستخدم جديد',  
                           form=form,  
                           active_page='add_user',  
                           active_menu='users')  

@admin_bp.route('/user/edit/<int:user_id>', methods=['GET', 'POST'])  
@login_required  
@admin_required  
def edit_user(user_id):  
    user = User.query.options(joinedload(User.student_profile)).get_or_404(user_id)  
    form = UserForm(obj=user, original_username=user.username, original_email=user.email)  

    # لتعبئة قائمة الطلاب غير المرتبطين في النموذج  
    unlinked_students = Student.query.filter(Student.user_account == None).order_by(Student.first_name).all()  
    form.student_to_link.choices = [(s.id, s.full_name) for s in unlinked_students]  
    form.student_to_link.choices.insert(0, ('', '--- اختر طالب للربط (اختياري) ---'))  

    # إذا كان المستخدم الحالي مرتبطًا بطالب، أضف هذا الطالب إلى الخيارات المتاحة كخيار افتراضي  
    if user.student_profile:  
        # تأكد من أن خيار الطالب الحالي موجود في القائمة  
        if (str(user.student_profile.id), user.student_profile.full_name) not in form.student_to_link.choices:  
            form.student_to_link.choices.insert(1, (str(user.student_profile.id), user.student_profile.full_name + ' (مرتبط حاليًا)'))  
        form.student_to_link.data = str(user.student_profile.id) # تعيين القيمة الافتراضية  

    if form.validate_on_submit():  
        user.username = form.username.data  
        user.email = form.email.data  
        user.role = form.role.data  
        if form.password.data:  
            user.set_password(form.password.data)  

        # ****** الجزء الجديد للربط عند تعديل المستخدم ******  
        if user.role == UserRole.STUDENT:  
            student_id_from_form = form.student_to_link.data  
            if student_id_from_form: # إذا تم اختيار طالب للربط  
                student_id_to_link = int(student_id_from_form)  
                student_to_assign = Student.query.get(student_id_to_link)  

                if student_to_assign:  
                    # إذا كان الطالب المراد ربطه هو نفسه الطالب المرتبط حاليًا، فلا تفعل شيئًا  
                    if user.student_profile and user.student_profile.id == student_to_assign.id:  
                        pass # لا تغيير في الربط  
                    # إذا كان الطالب المراد ربطه مرتبطًا بمستخدم آخر (غير المستخدم الحالي)  
                    elif student_to_assign.user_account and student_to_assign.user_account.id != user.id:  
                        flash(f'الطالب {student_to_assign.full_name} مرتبط بالفعل بالمستخدم {student_to_assign.user_account.username}.', 'danger')  
                        return render_template('admin/user_form.html', title=f'تعديل المستخدم: {user.username}', form=form, active_page='manage_users', active_menu='users')  
                    else:  
                        # ربط المستخدم بالطالب الجديد/المحدد  
                        user.student_profile = student_to_assign  
                else:  
                    flash('الطالب المحدد غير موجود.', 'danger')  
                    return render_template('admin/user_form.html', title=f'تعديل المستخدم: {user.username}', form=form, active_page='manage_users', active_menu='users')  
            else: # إذا لم يتم اختيار طالب للربط (أو تم إلغاء الربط)  
                user.student_profile = None # إزالة الربط  
        else: # إذا لم يكن الدور طالبًا، تأكد من عدم وجود ربط بطالب  
            user.student_profile = None  
        # **************************************************  

        db.session.commit()  
        flash('تم تحديث بيانات المستخدم بنجاح!', 'success')  
        return redirect(url_for('admin.manage_users'))  
    elif request.method == 'GET':  
        # تهيئة الفورم للـ GET request  
        if user.student_profile:  
            form.student_to_link.data = str(user.student_profile.id) # تعيين القيمة الافتراضية للطالب المرتبط  

    return render_template('admin/user_form.html',  
                           title=f'تعديل المستخدم: {user.username}',  
                           form=form,  
                           active_page='manage_users',  
                           active_menu='users')  
    
from app.models.notification import Notification
from app.models.excuse import Excuse
from app.models.attendance_log import AttendanceLog
    
@admin_bp.route('/user/delete/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    
    if user == current_user:
        flash('لا يمكنك حذف حسابك الحالي!', 'danger')
        return redirect(url_for('admin.manage_users'))
    
    try:
        # 🟢 1. احذف الإشعارات المرتبطة بهذا المستخدم أولاً
        # احذف الإشعارات اللي هو المرسل (sender)
        Notification.query.filter_by(sender_id=user_id).delete()
        
        # احذف الإشعارات اللي هو المستقبل (receiver)
        Notification.query.filter_by(receiver_id=user_id).delete()
        
        # 🟢 2. إذا كان المستخدم مرتبطًا بطالب
        if user.student_profile:
            student_id = user.student_profile.id
            
            # احذف الأعذار المرتبطة بالطالب
            Excuse.query.filter_by(student_id=student_id).delete()
            
            # احذف سجلات الحضور المرتبطة بالطالب
            AttendanceLog.query.filter_by(student_id=student_id).delete()
            
            # احذف بيانات الطالب نفسه
            user.student_profile = None
        
        # 🟢 3. احذف المستخدم نفسه
        db.session.delete(user)
        db.session.commit()
        
        flash('تم حذف المستخدم بنجاح!', 'success')
    
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف المستخدم: {str(e)}', 'danger')
        print(f"❌ Error deleting user: {e}")
    
    return redirect(url_for('admin.manage_users'))

# --- إدارة الطلاب (Students Management) ---  

@admin_bp.route('/students')  
@login_required  
@admin_required  
def manage_students():  
    students = Student.query.all()  
    return render_template('admin/manage_students.html',  
                           title='إدارة الطلاب',  
                           students=students,  
                           active_page='manage_students',  
                           active_menu='students')  

@admin_bp.route('/student/add', methods=['GET', 'POST'])  
@login_required  
@admin_required  
def add_student():  
    form = StudentForm()  
    # لتعبئة قائمة المستخدمين غير المرتبطين في النموذج  
    unlinked_users = User.query.filter(User.student_profile == None, User.role == UserRole.STUDENT).order_by(User.username).all()  
    form.user_to_link.choices = [(u.id, u.username) for u in unlinked_users]  
    form.user_to_link.choices.insert(0, ('', '--- اختر مستخدم للربط (اختياري) ---'))  

    if form.validate_on_submit():  
        student = Student(  
            student_id_number=form.student_id_number.data,  
            first_name=form.first_name.data,  
            last_name=form.last_name.data,  
            parent_email=form.parent_email.data,  
            parent_phone_number=form.parent_phone_number.data,  
            major=form.major.data,  
            grade=form.grade.data,  
            date_of_birth=form.date_of_birth.data,  
            is_active=form.is_active.data,  
            enrollment_date=datetime.now(pytz.utc),  
            created_at=datetime.now(pytz.utc),  
            updated_at=datetime.now(pytz.utc)  
        )  

        # ****** الجزء الجديد للربط عند إضافة الطالب ******  
        if form.user_to_link.data:  
            user_id_to_link = int(form.user_to_link.data)  
            user_to_assign = User.query.get(user_id_to_link)  

            if user_to_assign:  
                # تحقق إذا كان المستخدم مرتبطًا بالفعل بطالب آخر  
                if user_to_assign.student_profile:  
                    flash(f'المستخدم {user_to_assign.username} مرتبط بالفعل بالطالب {user_to_assign.student_profile.full_name}.', 'danger')  
                    return render_template('admin/student_form.html', title='إضافة طالب جديد', form=form, active_page='add_student', active_menu='students')  
                student.user_account = user_to_assign # ربط الطالب بالمستخدم  
            else:  
                flash('المستخدم المحدد غير موجود.', 'danger')  
                return render_template('admin/student_form.html', title='إضافة طالب جديد', form=form, active_page='add_student', active_menu='students')  
        # **************************************************  

        db.session.add(student)  
        db.session.commit()  
        flash('تم إضافة الطالب بنجاح!', 'success')  
        return redirect(url_for('admin.manage_students'))  
    return render_template('admin/student_form.html',  
                           title='إضافة طالب جديد',  
                           form=form,  
                           active_page='add_student',  
                           active_menu='students')  

@admin_bp.route('/student/edit/<int:id>', methods=['GET', 'POST'])  
@login_required  
@admin_required  
def edit_student(id):  
    student = Student.query.options(joinedload(Student.user_account)).get_or_404(id)  
    form = StudentForm(obj=student, original_student_id_number=student.student_id_number)  

    # لتعبئة قائمة المستخدمين غير المرتبطين في النموذج  
    unlinked_users = User.query.filter(User.student_profile == None, User.role == UserRole.STUDENT).order_by(User.username).all()  
    form.user_to_link.choices = [(u.id, u.username) for u in unlinked_users]  
    form.user_to_link.choices.insert(0, ('', '--- اختر مستخدم للربط (اختياري) ---'))  

    # إذا كان الطالب الحالي مرتبطًا بمستخدم، أضف هذا المستخدم إلى الخيارات المتاحة كخيار افتراضي  
    if student.user_account:  
        # تأكد من أن خيار المستخدم الحالي موجود في القائمة  
        if (str(student.user_account.id), student.user_account.username) not in form.user_to_link.choices:  
            form.user_to_link.choices.insert(1, (str(student.user_account.id), student.user_account.username + ' (مرتبط حاليًا)'))  
        form.user_to_link.data = str(student.user_account.id) # تعيين القيمة الافتراضية  

    if form.validate_on_submit():  
        student.student_id_number = form.student_id_number.data  
        student.first_name = form.first_name.data  
        student.last_name = form.last_name.data  
        student.parent_email = form.parent_email.data  
        student.parent_phone_number = form.parent_phone_number.data  
        student.major = form.major.data  
        student.grade = form.grade.data  
        student.date_of_birth = form.date_of_birth.data  
        student.is_active = form.is_active.data  
        student.updated_at = datetime.now(pytz.utc)  

        # ****** الجزء الجديد للربط عند تعديل الطالب ******  
        user_id_from_form = form.user_to_link.data  
        if user_id_from_form: # إذا تم اختيار مستخدم للربط  
            user_id_to_link = int(user_id_from_form)  
            user_to_assign = User.query.get(user_id_to_link)  

            if user_to_assign:  
                # إذا كان المستخدم المراد ربطه هو نفسه المستخدم المرتبط حاليًا، فلا تفعل شيئًا  
                if student.user_account and student.user_account.id == user_to_assign.id:  
                    pass # لا تغيير في الربط  
                # إذا كان المستخدم المراد ربطه مرتبطًا بطالب آخر (غير الطالب الحالي)  
                elif user_to_assign.student_profile and user_to_assign.student_profile.id != student.id:  
                    flash(f'المستخدم {user_to_assign.username} مرتبط بالفعل بالطالب {user_to_assign.student_profile.full_name}.', 'danger')  
                    return render_template('admin/student_form.html', title=f'تعديل الطالب: {student.full_name}', form=form, active_page='manage_students', active_menu='students')  
                else:  
                    # ربط الطالب بالمستخدم الجديد/المحدد  
                    student.user_account = user_to_assign  
            else:  
                flash('المستخدم المحدد غير موجود.', 'danger')  
                return render_template('admin/student_form.html', title=f'تعديل الطالب: {student.full_name}', form=form, active_page='manage_students', active_menu='students')  
        else: # إذا لم يتم اختيار مستخدم للربط (أو تم إلغاء الربط)  
            student.user_account = None # إزالة الربط  
        # **************************************************  

        db.session.commit()  
        flash('تم تحديث بيانات الطالب بنجاح!', 'success')  
        return redirect(url_for('admin.manage_students'))  
    elif request.method == 'GET':  
        # تهيئة الفورم للـ GET request  
        if student.user_account:  
            form.user_to_link.data = str(student.user_account.id) # تعيين القيمة الافتراضية للمستخدم المرتبط  

    return render_template('admin/student_form.html',  
                           title=f'تعديل الطالب: {student.full_name}',  
                           form=form,  
                           active_page='manage_students',  
                           active_menu='students')  

@admin_bp.route('/student/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_student(id):
    student = Student.query.get_or_404(id)
    
    try:
        # 🟢 1. احذف الإشعارات المرتبطة بهذا الطالب (عبر المستخدم)
        if student.user_account:
            user_id = student.user_account.id
            # احذف الإشعارات اللي الطالب هو المرسل (sender)
            Notification.query.filter_by(sender_id=user_id).delete()
            # احذف الإشعارات اللي الطالب هو المستقبل (receiver)
            Notification.query.filter_by(receiver_id=user_id).delete()
        
        # 🟢 2. احذف الأعذار المرتبطة بالطالب
        Excuse.query.filter_by(student_id=id).delete()
        
        # 🟢 3. احذف سجلات الحضور المرتبطة بالطالب
        AttendanceLog.query.filter_by(student_id=id).delete()
        
        # 🟢 4. فصل الربط مع المستخدم (إذا أردت حفظ بيانات المستخدم)
        if student.user_account:
            student.user_account = None
        
        # 🟢 5. احذف الطالب نفسه
        db.session.delete(student)
        db.session.commit()
        
        flash('تم حذف الطالب بنجاح!', 'success')
    
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف الطالب: {str(e)}', 'danger')
        print(f"❌ Error deleting student: {e}")
    
    return redirect(url_for('admin.manage_students'))

# --- إدارة البطاقات (Cards Management) ---  

@admin_bp.route('/cards')  
@login_required  
@admin_required  
def manage_cards():  
    cards = Card.query.all()  
    return render_template('admin/manage_cards.html',  
                           title='إدارة البطاقات',  
                           cards=cards,  
                           active_page='manage_cards',  
                           active_menu='cards')  

@admin_bp.route('/card/add', methods=['GET', 'POST'])  
@login_required  
@admin_required  
def add_card():  
    form = CardForm()  

    students = Student.query.order_by(Student.first_name).all()  
    form.student.choices = [(s.id, s.full_name) for s in students]  
    form.student.choices.insert(0, (0, '--- اختر طالب ---'))  

    if form.validate_on_submit():  
        existing_card = Card.query.filter_by(card_uid=form.card_uid.data).first()  
        if existing_card:  
            flash('رقم البطاقة هذا موجود بالفعل.', 'danger')  
            return render_template('admin/card_form.html',  
                                   title='إصدار بطاقة جديدة',  
                                   form=form,  
                                   active_page='add_card',  
                                   active_menu='cards')  

        student_id_to_assign = form.student.data if form.student.data != 0 else None  

        card = Card(card_uid=form.card_uid.data,  
                    issued_at=form.issued_at.data,  
                    status=form.status.data,  
                    student_id=student_id_to_assign)  
        db.session.add(card)  
        db.session.commit()  
        flash('تم إصدار البطاقة بنجاح!', 'success')  
        return redirect(url_for('admin.manage_cards'))  
    return render_template('admin/card_form.html',  
                           title='إصدار بطاقة جديدة',  
                           form=form,  
                           active_page='add_card',  
                           active_menu='cards')  
    
@admin_bp.route('/card/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_card(id):
    card = Card.query.get_or_404(id)
    form = CardForm()

    students = Student.query.order_by(Student.first_name).all()
    form.student.choices = [(s.id, s.full_name) for s in students]
    form.student.choices.insert(0, (0, '--- اختر طالب ---'))

    if form.validate_on_submit():
        # ✅ التحقق هنا في الـ routes بدل الـ forms
        # البحث عن بطاقة أخرى بنفس الـ UID
        existing_card = Card.query.filter(
            Card.card_uid == form.card_uid.data,
            Card.id != card.id  # استثنِ البطاقة الحالية
        ).first()
        
        if existing_card:
            flash('رقم البطاقة هذا موجود بالفعل لبطاقة أخرى.', 'danger')
            return render_template('admin/card_form.html',
                                   title=f'تعديل البطاقة: {card.card_uid}',
                                   form=form,
                                   active_page='manage_cards',
                                   active_menu='cards')

        # تحديث البيانات
        card.card_uid = form.card_uid.data
        card.issued_at = form.issued_at.data
        card.status = form.status.data
        card.student_id = form.student.data if form.student.data != 0 else None

        db.session.commit()
        flash('تم تحديث بيانات البطاقة بنجاح!', 'success')
        return redirect(url_for('admin.manage_cards'))
        
    elif request.method == 'GET':
        # ملء الفورم من قيم البطاقة الموجودة
        form.card_uid.data = card.card_uid
        form.student.data = card.student_id if card.student_id else 0
        if card.issued_at:
            form.issued_at.data = card.issued_at.date()
        form.status.data = card.status.value if card.status else None

    return render_template('admin/card_form.html',
                           title=f'تعديل البطاقة: {card.card_uid}',
                           form=form,
                           active_page='manage_cards',
                           active_menu='cards')

@admin_bp.route('/card/delete/<int:id>', methods=['POST'])  
@login_required  
@admin_required  
def delete_card(id):  
    card = Card.query.get_or_404(id)  
    db.session.delete(card)  
    db.session.commit()  
    flash('تم حذف البطاقة بنجاح!', 'success')  
    return redirect(url_for('admin.manage_cards'))  


# ... (بعد مسارات إدارة البطاقات) ...  

# ====================================================================  
# نقاط نهاية API الجديدة لعملية المسح  
# ====================================================================  

@admin_bp.route('/scan-card-for-form')  
@login_required  
@admin_required  
def scan_card_for_form():  
    """  
    Endpoint لانتظار قراءة بطاقة جديدة لإدخالها في النموذج  
    """  
    # الحصول على معرف الجلسة من الواجهة الأمامية  
    # هذا هو معرف الجلسة الذي ولده الـ JavaScript  
    client_session_id = request.headers.get('X-Session-ID')  

    if not client_session_id:  
        return jsonify({'error': 'معرف الجلسة مفقود في الرأس (X-Session-ID)'}), 400  

    timeout_seconds = 30  
    start_time = time.time()  

    with scan_lock:  
        global latest_scanned_card  
        # قم بتعيين session_id في latest_scanned_card إلى ما أرسلته الواجهة الأمامية  
        latest_scanned_card['session_id'] = client_session_id  
        latest_scanned_card['uid'] = None # إعادة تعيين UID لأي قراءة سابقة لنفس الجلسة  

    while (time.time() - start_time) < timeout_seconds:  
        with scan_lock:  
            # الآن نقارن session_id المخزن في latest_scanned_card  
            # مع session_id الذي أرسلته الواجهة الأمامية (client_session_id)  
            if (latest_scanned_card['uid'] is not None and  
                latest_scanned_card['session_id'] == client_session_id):  

                card_uid = latest_scanned_card['uid']  

                # مسح بيانات الجلسة بعد القراءة الناجحة  
                latest_scanned_card = {  
                    'uid': None,  
                    'timestamp': None,  
                    'session_id': None  
                }  

                return jsonify({  
                    'success': True,  
                    'card_uid': card_uid,  
                    'message': f'تم قراءة البطاقة: {card_uid}'  
                })  

        time.sleep(0.1)  

    with scan_lock:  
        # إذا انتهت المهلة، امسح بيانات الجلسة  
        if latest_scanned_card['session_id'] == client_session_id:  
            latest_scanned_card = {  
                'uid': None,  
                'timestamp': None,  
                'session_id': None  
            }  

    return jsonify({  
        'success': False,  
        'message': 'انتهت مدة الانتظار. لم يتم قراءة أي بطاقة.',  
        'timeout': True  
    })  

@admin_bp.route('/api/card-read', methods=['POST'])  
def handle_card_read():  
    """  
    Endpoint لاستقبال بيانات قراءة البطاقات من الأجهزة  
    """  
    try:  
        data = request.get_json()  
        card_uid = data.get('card_uid')  
        device_id = data.get('device_id')  
        # الأجهزة الخارجية قد تحتاج إلى إرسال session_id إذا كانت تعرفه  
        # أو يمكننا افتراض أن أي قراءة بطاقة أثناء جلسة مسح نشطة  
        # هي مخصصة لتلك الجلسة.  
        # للتبسيط، سنفترض أن أي قراءة بطاقة جديدة ستملأ الجلسة النشطة.  

        if not card_uid or not device_id:  
            return jsonify({'error': 'بيانات ناقصة (card_uid أو device_id)'}), 400  

        device = Device.query.filter_by(device_id=device_id, is_active=True).first()  
        if not device:  
            return jsonify({'error': 'جهاز غير موجود أو غير مفعل'}), 404  

        # لا تحتاج إلى البحث عن البطاقة هنا إذا كان الغرض هو فقط تحديث latest_scanned_card  
        # card = Card.query.filter_by(card_uid=card_uid).first()  
        # if not card:  
        #     pass  

        with scan_lock:  
            global latest_scanned_card  
            # نتحقق مما إذا كانت هناك جلسة مسح نشطة تنتظر  
            if latest_scanned_card['session_id'] is not None:  
                latest_scanned_card['uid'] = card_uid  
                latest_scanned_card['timestamp'] = datetime.now()  
            # else:  
            #     # إذا لم تكن هناك جلسة مسح نشطة، يمكننا معالجة هذا كحضور عادي  
            #     # أو تجاهله إذا كان هذا المسار مخصصًا فقط لتعبئة النموذج.  
            #     pass  

        # ... (يمكنك إضافة منطق تسجيل الحضور الفعلي هنا إذا لزم الأمر) ...  

        return jsonify({'success': True, 'message': 'تم معالجة البطاقة'})  

    except Exception as e:  
        print(f"Error in card read: {str(e)}")  
        return jsonify({'error': 'خطأ في معالجة البطاقة'}), 500  

@admin_bp.route('/cancel-card-scan', methods=['POST']) # تأكد من وجود methods=['POST']  
@login_required  
@admin_required  
def cancel_card_scan():  
    """  
    Endpoint لإلغاء عملية قراءة البطاقة  
    """  
    client_session_id = request.headers.get('X-Session-ID')  

    with scan_lock:  
        global latest_scanned_card  
        # نتحقق مما إذا كانت الجلسة التي يتم إلغاؤها هي الجلسة النشطة  
        if latest_scanned_card['session_id'] == client_session_id:  
            latest_scanned_card = {  
                'uid': None,  
                'timestamp': None,  
                'session_id': None  
            }  

    return jsonify({'success': True, 'message': 'تم إلغاء عملية القراءة'})  

# --- إدارة الأجهزة (Devices Management) ---  

@admin_bp.route('/devices')  
@login_required  
@admin_required  
def manage_devices():  
    devices = Device.query.all()  
    return render_template('admin/manage_devices.html',  
                           title='إدارة الأجهزة',  
                           devices=devices,  
                           active_page='manage_devices',  
                           active_menu='devices')  

@admin_bp.route('/device/add', methods=['GET', 'POST'])  
@login_required  
@admin_required  
def add_device():  
    form = DeviceForm()  
    if form.validate_on_submit():  
        device = Device(name=form.name.data,  
                        serial_number=form.serial_number.data,  
                        location=form.location.data,  
                        status=form.status.data)  
        db.session.add(device)  
        db.session.commit()  
        flash('تم إضافة الجهاز بنجاح!', 'success')  
        return redirect(url_for('admin.manage_devices'))  
    return render_template('admin/device_form.html',  
                           title='إضافة جهاز جديد',  
                           form=form,  
                           active_page='add_device',  
                           active_menu='devices')  

@admin_bp.route('/device/edit/<int:id>', methods=['GET', 'POST'])  
@login_required  
@admin_required  
def edit_device(id):  
    device = Device.query.get_or_404(id)  
    form = DeviceForm(obj=device, original_serial_number=device.serial_number)  
    if form.validate_on_submit():  
        device.name = form.name.data  
        device.serial_number = form.serial_number.data  
        device.location = form.location.data  
        device.status = form.status.data  
        db.session.commit()  
        flash('تم تحديث بيانات الجهاز بنجاح!', 'success')  
        return redirect(url_for('admin_bp.manage_devices'))  
    return render_template('admin/device_form.html',  
                           title=f'تعديل الجهاز: {device.name}',  
                           form=form,  
                           active_page='manage_devices',  
                           active_menu='devices')  

@admin_bp.route('/device/delete/<int:id>', methods=['POST'])  
@login_required  
@admin_required  
def delete_device(id):  
    device = Device.query.get_or_404(id)  
    db.session.delete(device)  
    db.session.commit()  
    flash('تم حذف الجهاز بنجاح!', 'success')  
    return redirect(url_for('admin_bp.manage_devices'))  

# --- مراقبة سجلات الحضور (Attendance Logs) ---  
SAUDIA_TZ = pytz.timezone('Asia/Riyadh')

@admin_bp.route('/attendance_logs')
@login_required
@admin_required # إذا كان هذا المسار للمشرفين فقط
def view_attendance_logs():
    attendance_logs_utc = AttendanceLog.query.options(
        joinedload(AttendanceLog.student),
        joinedload(AttendanceLog.device),
        joinedload(AttendanceLog.card)
    ).order_by(AttendanceLog.timestamp.desc()).all()

    # ***** هذا هو الجزء الجديد الذي يجب إضافته وتعديله *****
    processed_logs = []
    for log in attendance_logs_utc:
        # التأكد أن log.timestamp هو كائن aware (يحمل معلومات المنطقة الزمنية)
        # إذا كان مخزنًا كـ UTC (وهو الأفضل)، فسنقوم بتحويله
        if log.timestamp.tzinfo is None: # إذا كان naive (بدون معلومات منطقة زمنية)
            # نفترض أنه UTC لأنه تم تخزينه بـ datetime.now(pytz.utc)
            utc_dt = pytz.utc.localize(log.timestamp)
        else:
            utc_dt = log.timestamp # إذا كان aware بالفعل (مثل UTC)

        local_dt = utc_dt.astimezone(SAUDIA_TZ) # تحويله إلى التوقيت المحلي للسعودية

        # الطريقة الأبسط: تحديث خاصية timestamp في الكائن الحالي (لغرض العرض فقط)
        log.timestamp = local_dt
        processed_logs.append(log) # إضافة الكائن المعدّل إلى القائمة الجديدة

    return render_template('admin/attendance_logs.html',
                           title='سجلات الحضور',
                           logs=processed_logs, # ***** تمرير القائمة الجديدة التي تحتوي على التوقيتات المحلية *****
                           active_page='view_attendance_logs',
                           active_menu='attendance_logs')

# --- التقارير (Reports) ---  

@admin_bp.route('/reports')  
@login_required  
@admin_required  
def reports_dashboard():  
    return render_template('admin/reports_dashboard.html',  
                           title='لوحة تحكم التقارير',  
                           active_page='reports_dashboard',  
                           active_menu='reports')  

@admin_bp.route('/report/student/<int:student_id>/attendance')  
@login_required  
@admin_required  
def student_attendance_report(student_id):  
    student = Student.query.get_or_404(student_id)  

    attendance_logs = AttendanceLog.query.filter_by(student_id=student.id).order_by(AttendanceLog.timestamp.desc()).all()  

    return render_template('admin/student_attendance_report.html',  
                           title=f'تقرير حضور {student.first_name} {student.last_name}',  
                           student=student,  
                           attendance_logs=attendance_logs,  
                           active_page='student_attendance_report',  
                           active_menu='reports')  

@admin_bp.route('/report/card_status')  
@login_required  
@admin_required  
def card_status_report():  
    total_cards = Card.query.count()  
    active_cards_count = Card.query.filter_by(status='ACTIVE').count()  
    inactive_cards_count = Card.query.filter_by(status='INACTIVE').count()  
    lost_cards_count = Card.query.filter_by(status='LOST').count()  

    all_cards = Card.query.options(db.joinedload(Card.student)).all()  

    return render_template('admin/card_status_report.html',  
                           title='تقرير حالة البطاقات',  
                           total_cards=total_cards,  
                           active_cards_count=active_cards_count,  
                           inactive_cards_count=inactive_cards_count,  
                           lost_cards_count=lost_cards_count,  
                           all_cards=all_cards,  
                           active_page='card_status_report',  
                           active_menu='reports')  


@admin_bp.route('/report/device_activity')  
@login_required  
@admin_required  
def generate_device_activity_report():  
    return render_template('admin/device_activity_report.html',  
                           title='تقرير نشاط الأجهزة',  
                           active_page='device_activity_report',  
                           active_menu='reports')  

@admin_bp.route('/student/<int:student_id>/profile')  
@login_required  
@admin_required  
def student_profile(student_id):  
    student = Student.query.get_or_404(student_id)  
    return render_template('admin/student_profile.html', student=student, title=f'ملف الطالب: {student.first_name} {student.last_name}')  


@admin_bp.route('/students_list')  
@login_required  
@admin_required  
def students_list():  
    students = Student.query.all()  
    return render_template('admin/students_list.html', students=students, title='قائمة الطلاب')  

@admin_bp.route('/report/system_summary')  
@login_required  
@admin_required  
def generate_system_summary_report():  
    total_users = User.query.count()  
    total_students = Student.query.count()  
    total_cards = Card.query.count()  
    total_devices = Device.query.count()  
    total_attendance_logs = AttendanceLog.query.count()  

    return render_template('admin/system_summary_report.html',  
                           title='تقرير ملخص النظام',  
                           total_users=total_users,  
                           total_students=total_students,  
                           total_cards=total_cards,  
                           total_devices=total_devices,  
                           total_attendance_logs=total_attendance_logs,  
                           active_page='system_summary_report',  
                           active_menu='reports')  

@admin_bp.route('/settings')  
@login_required  
@admin_required  
def system_settings():  
    return render_template('admin/system_settings.html',  
                           title='إعدادات النظام',  
                           active_page='system_settings',  
                           active_menu='settings')
    
    
# تحديث فوري للصفحهبدون اعادة تحميل لها
@admin_bp.route('/api/attendance_logs')
@login_required
@admin_required
def api_attendance_logs():
    """
    API endpoint لتحديث الحضور الفعلي
    يرجع آخر السجلات بصيغة JSON
    """
    try:
        # جلب آخر 50 سجل مع تحميل العلاقات
        attendance_logs = AttendanceLog.query.options(
            joinedload(AttendanceLog.student),
            joinedload(AttendanceLog.device),
            joinedload(AttendanceLog.card)
        ).order_by(AttendanceLog.timestamp.desc()).limit(50).all()

        # المنطقة الزمنية
        SAUDIA_TZ = pytz.timezone('Asia/Riyadh')

        # تحويل البيانات إلى JSON
        logs_data = []
        for log in attendance_logs:
            # تحويل التوقيت من UTC إلى التوقيت المحلي
            if log.timestamp.tzinfo is None:
                utc_dt = pytz.utc.localize(log.timestamp)
            else:
                utc_dt = log.timestamp
            
            local_dt = utc_dt.astimezone(SAUDIA_TZ)

            # ✅ التأكد من البيانات الموجودة
            student_data = log.student
            device_data = log.device
            card_data = log.card

            logs_data.append({
                'id': log.id,
                # ✅ استخدام student_id_number بدل None
                'student_id_number': student_data.student_id_number if student_data else 'N/A',
                'student_name': student_data.full_name if student_data else 'غير معروف',
                'device_name': device_data.name if device_data else 'غير معروف',
                'device_location': device_data.location if device_data else 'غير متوفر',
                # ✅ استخدام التوقيت المحلي (Riyadh)
                'timestamp': local_dt.isoformat(),
                'status': log.status.name if log.status else 'UNKNOWN',
                'card_uid': card_data.card_uid if card_data else None
            })

        return jsonify({
            'status': 'success',
            'logs': logs_data,
            'count': len(logs_data)
        }), 200

    except Exception as e:
        current_app.logger.error(f"Error in api_attendance_logs: {str(e)}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

# ====================================================================
# Endpoints التقارير والاستخراج
# ====================================================================

@admin_bp.route('/api/export-attendance-report', methods=['POST'])
@login_required
@admin_required
def export_attendance_report():
    """استخراج سجلات الحضور بصيغ مختلفة"""
    try:
        export_format = request.form.get('format', 'csv')
        date_from = request.form.get('date_from')
        date_to = request.form.get('date_to')
        student_id = request.form.get('student_id')
        device_id = request.form.get('device_id')
        
        # بناء الـ query
        query = AttendanceLog.query.options(
            joinedload(AttendanceLog.student),
            joinedload(AttendanceLog.device)
        )
        
        # التصفية حسب التاريخ
        if date_from:
            start_date = dt.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(db.func.date(AttendanceLog.timestamp) >= start_date)
        
        if date_to:
            end_date = dt.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(db.func.date(AttendanceLog.timestamp) <= end_date)
        
        # التصفية حسب الطالب
        if student_id:
            query = query.filter(AttendanceLog.student_id == student_id)
        
        # التصفية حسب الجهاز
        if device_id:
            query = query.filter(AttendanceLog.device_id == device_id)
        
        logs = query.order_by(AttendanceLog.timestamp.desc()).all()
        
        if export_format == 'excel':
            return export_to_excel(logs)
        elif export_format == 'csv':
            return export_to_csv(logs)
        elif export_format == 'pdf':
            return export_to_pdf(logs)
        else:
            return jsonify({'success': False, 'message': 'صيغة غير مدعومة'}), 400
            
    except Exception as e:
        current_app.logger.error(f"Error exporting report: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


def export_to_csv(logs):
    """تصدير إلى CSV"""
    try:
        output = io.StringIO()
        writer = csv.writer(output, encoding='utf-8')
        
        # الرؤوس
        writer.writerow(['#', 'اسم الطالب', 'الرقم الأكاديمي', 'الجهاز', 'الموقع', 'التاريخ والوقت', 'الحالة'])
        
        # البيانات
        for index, log in enumerate(logs, 1):
            writer.writerow([
                index,
                log.student.full_name if log.student else 'غير معروف',
                log.student.student_id_number if log.student else 'N/A',
                log.device.name if log.device else 'غير معروف',
                log.device.location if log.device else 'N/A',
                log.timestamp.strftime('%Y-%m-%d %I:%M:%S %p'),
                log.status.value if log.status else 'N/A'
            ])
        
        # إرسال الملف
        output.seek(0)
        return output.getvalue(), 200, {
            'Content-Type': 'text/csv; charset=utf-8',
            'Content-Disposition': f'attachment; filename="attendance_{dt.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        }
    except Exception as e:
        current_app.logger.error(f"Error in CSV export: {str(e)}")
        raise


def export_to_excel(logs):
    """تصدير إلى Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "سجلات الحضور"
        
        # الرؤوس
        headers = ['#', 'اسم الطالب', 'الرقم الأكاديمي', 'الجهاز', 'الموقع', 'التاريخ والوقت', 'الحالة']
        ws.append(headers)
        
        # تنسيق الرؤوس
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # البيانات
        for index, log in enumerate(logs, 1):
            ws.append([
                index,
                log.student.full_name if log.student else 'غير معروف',
                log.student.student_id_number if log.student else 'N/A',
                log.device.name if log.device else 'غير معروف',
                log.device.location if log.device else 'N/A',
                log.timestamp.strftime('%Y-%m-%d %I:%M:%S %p'),
                'دخول' if log.status.value == 'ENTER' else 'خروج' if log.status.value == 'EXIT' else log.status.value
            ])
        
        # ضبط عرض الأعمدة
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 10
        
        # حفظ في الذاكرة
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue(), 200, {
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': f'attachment; filename="attendance_{dt.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        }
    except ImportError:
        # إذا لم تكن openpyxl مثبتة، قم بإرجاع خطأ
        return jsonify({'success': False, 'message': 'مكتبة openpyxl غير مثبتة. الرجاء تثبيتها: pip install openpyxl'}), 500


def export_to_pdf(logs):
    """تصدير إلى PDF"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        
        # العنوان
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=12,
            alignment=1  # مركز
        )
        title = Paragraph('تقرير سجلات الحضور', title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.3 * inch))
        
        # البيانات في جدول
        table_data = [['#', 'اسم الطالب', 'الرقم الأكاديمي', 'الجهاز', 'التاريخ والوقت', 'الحالة']]
        
        for index, log in enumerate(logs[:100], 1):  # أول 100 سجل فقط
            table_data.append([
                str(index),
                log.student.full_name if log.student else 'غير معروف',
                log.student.student_id_number if log.student else 'N/A',
                log.device.name if log.device else 'غير معروف',
                log.timestamp.strftime('%Y-%m-%d %I:%M:%S %p'),
                'دخول' if log.status.value == 'ENTER' else 'خروج'
            ])
        
        table = Table(table_data, colWidths=[0.5*inch, 1.5*inch, 1.2*inch, 1.2*inch, 1.3*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        # بناء الـ PDF
        doc.build(elements)
        output.seek(0)
        
        return output.getvalue(), 200, {
            'Content-Type': 'application/pdf',
            'Content-Disposition': f'attachment; filename="attendance_{dt.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        }
    except Exception as e:
        current_app.logger.error(f"Error in PDF export: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_bp.route('/api/today-statistics')
@login_required
@admin_required
def today_statistics():
    """الحصول على إحصائيات اليوم"""
    try:
        today = datetime.now(pytz.utc).date()
        
        # السجلات اليومية
        today_logs = AttendanceLog.query.filter(
            db.func.date(AttendanceLog.timestamp) == today
        ).all()
        
        # حساب الدخول والخروج
        present_count = sum(1 for log in today_logs if log.status.value == 'ENTER')
        absent_count = sum(1 for log in today_logs if log.status.value == 'EXIT')
        
        return jsonify({
            'success': True,
            'present_count': present_count,
            'absent_count': absent_count,
            'total_logs': len(today_logs)
        }), 200
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_bp.route('/api/generate-custom-report', methods=['POST'])
@login_required
@admin_required
def generate_custom_report():
    """إنشاء تقرير مخصص"""
    try:
        data = request.get_json()
        date_from = data.get('date_from')
        date_to = data.get('date_to')
        student_id = data.get('student_id')
        device_id = data.get('device_id')
        
        query = AttendanceLog.query.options(
            joinedload(AttendanceLog.student),
            joinedload(AttendanceLog.device)
        )
        
        if date_from:
            start_date = dt.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(db.func.date(AttendanceLog.timestamp) >= start_date)
        
        if date_to:
            end_date = dt.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(db.func.date(AttendanceLog.timestamp) <= end_date)
        
        if student_id:
            query = query.filter(AttendanceLog.student_id == student_id)
        
        if device_id:
            query = query.filter(AttendanceLog.device_id == device_id)
        
        logs = query.order_by(AttendanceLog.timestamp.desc()).all()
        
        return jsonify({
            'success': True,
            'message': f'تم إنشاء تقرير مخصص يحتوي على {len(logs)} سجل',
            'count': len(logs)
        }), 200
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500        
    

# ... (الـ imports الأخرى مثل AttendanceLog, Student, Device, etc.) ...

@admin_bp.route('/generate_attendance_report', methods=['GET', 'POST'])
@login_required
@admin_required
def generate_attendance_report():
    if request.method == 'POST':
        report_date_str = request.form.get('report_date')
        start_time_str = request.form.get('start_time')
        end_time_str = request.form.get('end_time')
        location = request.form.get('location')

        if not all([report_date_str, start_time_str, end_time_str, location]):
            flash('جميع حقول التقرير مطلوبة!', 'danger')
            return redirect(url_for('admin.generate_attendance_report')) # العودة لصفحة النموذج

        try:
            report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()
            start_time = datetime.strptime(start_time_str, '%H:%M').time()
            end_time = datetime.strptime(end_time_str, '%H:%M').time()
        except ValueError:
            flash('تنسيق التاريخ أو الوقت غير صحيح.', 'danger')
            return redirect(url_for('admin.generate_attendance_report')) # العودة لصفحة النموذج

        # ** فترة السماح للتأخير (10 دقائق) **
        LATE_ALLOWANCE_MINUTES = 10 
        
        # تحويل start_time إلى datetime object في المنطقة الزمنية السعودية (لتسهيل حساب الفروقات)
        report_datetime_start = datetime.combine(report_date, start_time)
        report_datetime_end = datetime.combine(report_date, end_time)
        
        # قم بتوطين أوقات التقرير إلى SAUDIA_TZ ثم تحويلها إلى UTC للمقارنة مع DB
        # هذا ضروري لأن التوقيتات في DB (AttendanceLog.timestamp) يُفترض أنها UTC
        report_datetime_start_saudia_aware = SAUDIA_TZ.localize(report_datetime_start)
        report_datetime_end_saudia_aware = SAUDIA_TZ.localize(report_datetime_end)

        report_datetime_start_utc = report_datetime_start_saudia_aware.astimezone(pytz.utc)
        report_datetime_end_utc = report_datetime_end_saudia_aware.astimezone(pytz.utc)

        # حساب وقت انتهاء فترة السماح (بالتوقيت المحلي ثم تحويله لـ UTC للمقارنة إذا لزم الأمر)
        late_cutoff_local_time = (report_datetime_start + timedelta(minutes=LATE_ALLOWANCE_MINUTES)).time()
        
        # إذا كان وقت الانتهاء أقل من وقت البدء، هذا يعني أنه يمتد لليوم التالي
        # لكن لمنع التعقيد، سنفترض أن وقت الانتهاء لا يكون قبل وقت البدء لنفس اليوم في هذه المرحلة
        if end_time < start_time:
            flash("وقت الانتهاء لا يمكن أن يكون قبل وقت البدء لنفس اليوم.", "warning")
            return redirect(url_for('admin.generate_attendance_report'))


        # جلب جميع الطلاب
        all_students = Student.query.all()
        report_results = []

        for student in all_students:
            # البحث عن أول سجل دخول (ENTER) للطالب في الموقع المحدد والفترة
            first_entry_log = db.session.query(AttendanceLog)\
                                    .filter(
                                        AttendanceLog.student_id == student.id,
                                        AttendanceLog.device.has(location=location), # البحث بناءً على موقع الجهاز
                                        AttendanceLog.timestamp >= report_datetime_start_utc,
                                        AttendanceLog.timestamp <= report_datetime_end_utc, # البحث حتى نهاية الفترة
                                        AttendanceLog.status == AttendanceStatus.ENTER # فقط سجلات الدخول
                                    )\
                                    .order_by(AttendanceLog.timestamp.asc())\
                                    .first() # جلب أقدم سجل دخول

            status = 'غائب' # الافتراضي
            actual_entry_time = None
            
            if first_entry_log:
                entry_time_utc = first_entry_log.timestamp
                
                # 🟢 التعديل الحاسم هنا: تأكد أن التوقيت من DB هو "aware" كـ UTC
                if entry_time_utc.tzinfo is None:
                    # إذا كان "naive" (بدون معلومات المنطقة الزمنية)، افترضه UTC
                    entry_time_utc = pytz.utc.localize(entry_time_utc)
                
                # تحويل وقت الدخول الفعلي من UTC إلى توقيت السعودية للمقارنة والعرض
                entry_time_saudia_aware = entry_time_utc.astimezone(SAUDIA_TZ)
                actual_entry_time = entry_time_saudia_aware.time() # الوقت فقط بتوقيت الرياض
                
                # مقارنة وقت الدخول الفعلي مع أوقات التقرير (بالتوقيت المحلي)
                # هذه المقارنات (actual_entry_time <= start_time) صحيحة الآن
                # لأن start_time و late_cutoff_local_time هما أيضاً بتوقيت الرياض المحلي
                if actual_entry_time <= start_time:
                    status = 'حاضر'
                elif start_time < actual_entry_time <= late_cutoff_local_time:
                    status = 'متأخر'
                else: # دخل بعد فترة السماح
                    status = 'غائب (متأخر جداً)' # يمكنك تسميتها هكذا أو مجرد 'غائب'
            
            report_results.append({
                'student': student,
                'status': status,
                'actual_entry_time': actual_entry_time
            })
        
        # تخزين معايير التقرير في الجلسة لاستخدامها عند "تعميم التقرير"
        session['report_criteria'] = {
            'report_date': report_date_str,
            'start_time': start_time_str,
            # ... (بقية الكود) ...
            'end_time': end_time_str,
            'location': location,
            'late_allowance_minutes': LATE_ALLOWANCE_MINUTES # قد تحتاجها لاحقاً
        }
        
        # تخزين النتائج المؤقتة في الجلسة لاستخدامها عند "تعميم التقرير"
        session['report_temporary_results'] = [
            {'student_id': r['student'].id, 'status': r['status'], 'actual_entry_time': str(r['actual_entry_time']) if r['actual_entry_time'] else None}
            for r in report_results
        ]


        return render_template('admin/attendance_report_view.html', 
                               report_date=report_date.strftime('%Y-%m-%d'),
                               start_time=start_time_str,
                               end_time=end_time_str,
                               location=location,
                               report_results=report_results)
    else:
        # عرض نموذج توليد التقرير (GET request)
        today_date = date.today().strftime('%Y-%m-%d')
        # جلب جميع المواقع الفريدة من الأجهزة
        locations = db.session.query(Device.location).distinct().all()
        locations = [loc[0] for loc in locations if loc[0]] # استخراج اسم الموقع فقط وتصفية None
        return render_template('admin/generate_report_form.html', today_date=today_date, locations=locations)
    
    
    
@admin_bp.route('/admin/finalize_attendance_report')
@login_required
@admin_required
def finalize_attendance_report():
    report_criteria = session.get('report_criteria')
    report_temporary_results = session.get('report_temporary_results')

    if not report_criteria or not report_temporary_results:
        flash('لا توجد نتائج تقرير مؤقتة لتعميمها. يرجى إنشاء التقرير أولاً.', 'warning')
        return redirect(url_for('admin.generate_attendance_report'))

    try:
        report_date = datetime.strptime(report_criteria['report_date'], '%Y-%m-%d').date()
        start_time = datetime.strptime(report_criteria['start_time'], '%H:%M').time()
        end_time = datetime.strptime(report_criteria['end_time'], '%H:%M').time()
        location = report_criteria['location']

        for result in report_temporary_results:
            student_id = result['student_id']
            status = result['status']
            # قد يكون actual_entry_time_str None إذا كان الطالب غائباً
            actual_entry_time_str = result['actual_entry_time']
            actual_entry_time = datetime.strptime(actual_entry_time_str, '%H:%M:%S').time() if actual_entry_time_str else None

            # تحقق مما إذا كان هناك سجل موجود بالفعل لهذه الفترة لمنع التكرار
            existing_summary = AttendanceSummary.query.filter_by(
                student_id=student_id,
                report_date=report_date,
                start_time=start_time,
                location=location
            ).first()

            if existing_summary:
                # تحديث السجل الموجود
                existing_summary.status = status
                existing_summary.actual_entry_time = actual_entry_time
                db.session.add(existing_summary)
            else:
                # إنشاء سجل جديد
                new_summary = AttendanceSummary(
                    student_id=student_id,
                    report_date=report_date,
                    start_time=start_time,
                    end_time=end_time,
                    location=location,
                    status=status,
                    actual_entry_time=actual_entry_time
                )
                db.session.add(new_summary)

        db.session.commit()
        flash('تم تعميم تقرير الحضور بنجاح وحفظه في السجلات الدائمة!', 'success')
        
        # مسح النتائج المؤقتة ومعايير التقرير من الجلسة بعد التعميم
        session.pop('report_criteria', None)
        session.pop('report_temporary_results', None)

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error finalizing attendance report: {e}", exc_info=True)
        flash(f'حدث خطأ أثناء تعميم التقرير. يرجى التحقق من السجلات.', 'danger')
    
    return redirect(url_for('admin.generate_attendance_report'))    


@admin_bp.route('/profile') # <-- هذا هو المسار الجديد
@login_required
@admin_required # بما أنه في blueprint الأدمن، فمن المنطقي أن يكون للأدمن فقط
def profile(): # <-- هذا هو اسم الدالة، وبالتالي اسم الـ endpoint الجزئي
    """
    يعرض صفحة الملف الشخصي للمستخدم الحالي (الأدمن).
    """
    return render_template('admin/profile.html', # ستحتاج لإنشاء هذا القالب في الخطوة التالية
                           title='ملفي الشخصي',
                           user=current_user, # current_user يوفرها Flask-Login
                           active_page='profile', # يمكنك استخدامها لتحديد العنصر النشط في القائمة
                           active_menu='profile')
